import ctypes
import json
import openpyxl as xl
from datetime import datetime


def suppress_null(val):
    if not val:
        return 0
    else:
        return val


class VolCov:
    """
    A class used to get data from a volume calculator file and format it to a Diagram Class
    """

    def __init__(self, VC_file_directory=""):
        """ The constructor of the Volume calculator format. To initialize, gets the directory of the VC file and
        sets each sublist to the right property. """

        self.__VC_file = VC_file_directory
        self.vc_json = {}

    def toJSON(self):
        print("here")
        VC_list, xl_prop = self.getVC()
        VC_titles = ["Morning_Volumes", "Regular_Arrows", "General_Information", "LRT_Information",
                     "PublicTransport_Arrows", "Evening_Volumes", "Street_Names"]
        VC_id_titles = ["Project_Name", "Project_Number", "Project_Author", "Project_Count", "Project_Info"]
        id_info = VC_list.pop(-1)
        print(id_info)
        print(VC_list)
        print("-----general-----")
        id_json = {}
        for title, data in zip(VC_titles, VC_list):
            self.vc_json[title] = data
            print(title, ":", data)
        print("-----id-info-----")
        for id_title, id_data in zip(VC_id_titles, id_info):
            print(id_title, ":", id_data)
            if id_title == "Project_Author":
                id_json["Project_Author"] = id_data
            else:
                id_json[id_title] = id_data

        self.vc_json["ID_Information"] = id_json

        print(self.vc_json)
        return self.vc_json

    def saveFile(self, directory=""):
        time_stamp = datetime.now().strftime("%Y_%m_%d-%I_%M_%S_%p")
        location = directory + "×JUCSON×" + time_stamp + ".json"
        with open(location, "w+") as f:
            json.dump(self.toJSON(), f)

    def read_from_excel(self, run):
        wb = xl.load_workbook(self.VC, data_only=True)
        ws = wb.active
        volume = []
        lanes = []
        public_trans = []
        junc_public_trans = []
        streets = []
        junc_instructions = [suppress_null(ws.cell(row=36 + i, column=19).value) for i in range(5)]

        for i in range(4):
            temp_volume = [suppress_null(ws.cell(row=4 + run, column=4 + 4 * i + j).value) for j in range(3)]
            volume += temp_volume
            temp_lanes = [suppress_null(ws.cell(row=8, column=3 + 8 * i + j).value) for j in range(7)]
            lanes += temp_lanes
            temp_public_trans = [suppress_null(ws.cell(row=9, column=3 + 8 * i + j).value) for j in range(7)]
            public_trans += temp_public_trans
            temp_junc_public_trans = [suppress_null(ws.cell(row=9, column=3 + 8 * i + j).value) for j in range(7)]
            junc_public_trans += temp_junc_public_trans

        instructions = [suppress_null(ws.cell(row=36 + i, column=22).value) for i in range(11)]
        for i in range(12):
            try:
                volume[i] = round(volume[i] * instructions[10], 0)
            except:
                error = "volume must be put as numbers"
                MessageBox = ctypes.windll.user32.MessageBoxW
                MessageBox(None, error, 'Phaser error', 0)
                exit()
        rakal_instructions = [suppress_null(ws.cell(row=36 + i, column=26).value) for i in range(6)]

        for i in range(6):
            if i == 4 and rakal_instructions[i] == 1.125:
                i = i + 1
            if not isinstance(rakal_instructions[i], int):
                error = "rakal instructions table must be integers"
                MessageBox = ctypes.windll.user32.MessageBoxW
                MessageBox(None, error, 'Phaser error', 0)
                exit()
        for i in range(11):

            if i == 10 and isinstance(instructions[i], float):
                break
            if not isinstance(instructions[i], int):
                error = "instructions table must be integers"
                MessageBox = ctypes.windll.user32.MessageBoxW
                MessageBox(None, error, 'Phaser error', 0)
                exit()

        for rakal_cell in range(12):
            if (rakal_cell + 2) % 3 != 0:
                volume[rakal_cell] = round(volume[rakal_cell] * rakal_instructions[4], 0)
        if instructions[7] == 1:
            for m in range(28):
                public_trans[m] = 0
        for cell_street in range(4):
            temp_street = [suppress_null(ws.cell(row=4, column=22 + cell_street).value)]
            if temp_street == 0:
                temp_street = ""
            streets += temp_street

        prop = wb.properties
        return volume, lanes, instructions, rakal_instructions, public_trans, streets, \
               junc_instructions, wb.properties, junc_public_trans

    def getVC(self):
        run = 0
        junc_list = []
        excel_properties_list = []

        while run < 2:
            wb = xl.load_workbook(self.VC)

            active = wb.active
            volume, lanes, instructions, rakal_instructions, public_trans, streets, junc_instructions, \
            excel_properties, junc_public_trans = self.read_from_excel(run)

            junc_list.append(volume)

            if run == 0:
                junc_list.append(lanes)
                junc_list.append(instructions)
                junc_list.append(rakal_instructions)
                junc_list.append(junc_public_trans)

            if run == 1:
                junc_list.append(streets)
                junc_list.append(junc_instructions)
                excel_properties_list.append(excel_properties)

            run = run + 1

        return junc_list, excel_properties_list

    @property
    def VC(self):
        """Get the VC file full info"""
        return self.__VC_file

    @VC.setter
    def VC(self, vc_file):
        """Set the VC file full info"""
        self.__VC_file = vc_file
